from pprint import pprint as pp
from pathlib import Path
import re
import sys
from unittest import result
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

# UECapabilityCheck Tool.

# Global variable.
feat_set_dict = {}


def get_braces_array(content, indent):
    ret_lst = []

    for m in re.finditer("\n {{{0}}}{{".format(indent), content):
        work = content[m.start():]
        picked = ""
        pos = 0
        for each in work:
            picked += each
            if each == '{':
                pos += 1
            elif each == '}':
                pos -= 1
                if pos == 0:
                    ret_lst.append(picked)
                    break

    return ret_lst


def get_braces(content, target):
    m = re.search(target, content)
    if not m:
        print("get_braces: ", target + " is not found")
        return "", 0

    before_find = content[0:m.start()]

    work = content[m.end():]
    if work.index('{') == -1 or work.index('}') == -1 or \
            work.index('{') > work.index('}'):
        print(target + ": braces is not found or { is not found before }")
        return "", 0

    picked = ""
    pos = 0
    for each in work:
        picked += each
        if each == '{':
            pos += 1
        elif each == '}':
            pos -= 1
            if pos == 0:
                break
    else:
        print(target + " with braces does not match")
        return "", 0

    # if first line does not contain "{", remove it.
    picked_lst = picked.splitlines()
    if picked_lst[0].count("{") == 0:
        picked = "\n".join(picked_lst[1:])

    return picked, before_find.count("\n") + 1


def get_num(content, target):
    m = re.search(f"{target} +([0-9]+)", content)
    if m:
        return int(m.group(1))
    else:
        print(f"{target} is not found, return 0")
        return 0


def delete_indent(target, indent_num):
    work_lst = []
    for line in target.splitlines():
        work_lst.append(re.sub(f"^ {{{indent_num}}}", "", line))

    return "\n".join(work_lst)


def pick_inside_brace(input_str):
    # //Only pick "pick there" with deleting indent.
    # //   {
    # //    pick there
    # //    }

    ret_lst = []
    find_brace = False
    for line in input_str.split("\n"):
        if "}" in line:
            break
        if find_brace:
            ret_lst.append(re.sub("^ +", "", line))
        elif "{" in line:
            find_brace = True

    return "\n".join(ret_lst)


def get_supportedBandListEUTRA(log, result_dict):
    picked, find_line = get_braces(log, "supportedBandListEUTRA\n")
    int_lst = []

    if find_line > 0:
        for each_num in re.findall('([0-9]+)', picked):
            int_lst.append(int(each_num))

        int_lst.sort()
        result_dict["supportedBandListEUTRA"] = [str(a) for a in int_lst]


def get_supportedBandListUTRA(log, result_dict):
    picked, find_line = get_braces(log, "supportedBandListUTRA-FDD\n")
    int_list = []

    if find_line > 0:
        if "{" in picked:
            picked = picked.replace("{", "")
        if "}" in picked:
            picked = picked.replace("}", "")
        int_list =  picked.replace("\n", "").split(",")
        result_dict["supportedBandListUTRA-FDD"] = [str(a.strip()) for a in int_list]


def ca_combo_parse(block, indent):
    band = 0
    ul_lst = []
    dl_lst = []
    ret = {"ca": "", "ca_ul": "", "ca_layer": "", "log_snip": ""}

    for each_band in get_braces_array(block, indent):
        dl_class = ""

        m = re.search("bandEUTRA-r1[0-9] ([0-9]+)", each_band)
        if m:
            band = int(m.group(1))
        else:
            print("No proper band cannot be taken, Exit.")
            sys.exit()

        m = re.search("ca-BandwidthClassUL-r1[0-9] +([a-zA-Z]+)", each_band)
        if m:
            class_u = m.group(1).upper()
            ul_lst.append([band, class_u])

        m = re.search("ca-BandwidthClassDL-r1[0-9] +([a-zA-Z]+)", each_band)
        if m:
            dl_class = m.group(1).upper()

        m = re.search("supportedMIMO-CapabilityDL-r1[0-9] ([a-zA-Z]+)", each_band)
        if m:
            layer = m.group(1)
            layer = re.sub("twoLayers", "2", layer)
            layer = re.sub("fourLayers", "4", layer)
            layer_lst = [layer]

            if dl_class == "B" or dl_class == "C":
                layer_lst = layer_lst * 2
            if dl_class == "D":
                layer_lst = layer_lst * 3
            if dl_class == "E":
                layer_lst = layer_lst * 4
            if dl_class == "F":
                layer_lst = layer_lst * 5

            dl_lst.append([band, dl_class, "+".join(layer_lst)])

    if len(dl_lst) > 0:
        dl_lst.sort()
        dl_str = "CA_"
        if len(dl_lst) == 1 and dl_lst[0][1] == "A":
            dl_str = ""
        ret["ca"] = f"{dl_str}{'-'.join([str(a[0]) + a[1] for a in dl_lst])}"
        ret["ca_layer"] = f"{'+'.join([a[2] for a in dl_lst])}"

    if len(ul_lst) > 0:
        ul_lst.sort()
        up_str = "ULCA_"
        if len(ul_lst) == 1 and ul_lst[0][-1] == "A":
            up_str = ""
        ret["ca_ul"] = f"{up_str}{'-'.join([str(a[0]) + a[1] for a in ul_lst])}"

    log_snip = block
    ret["log_snip"] = "\n".join(log_snip.split("\n")[1:])

    return ret


def get_supportedBandCombination(log, result_dict):
    picked, find_line = get_braces(log, "supportedBandCombination-r10\n")
    if find_line > 0:
        for each_ca in get_braces_array(picked, 12):
            result_dict["CA"].append(ca_combo_parse(each_ca, 14))

    picked, find_line = get_braces(log, "supportedBandCombinationAdd-r11\n")
    if find_line > 0:
        for each_ca in get_braces_array(picked, 22):
            picked_ca, find_line2 = get_braces(each_ca, "bandParameterList-r11")
            if find_line2 > 0:
                result_dict["CA"].append(ca_combo_parse(picked_ca, 26))

    picked, find_line = get_braces(log, "supportedBandCombinationReduced-r13\n")
    if find_line > 0:
        for each_ca in get_braces_array(picked, 34):
            picked_ca, find_line2 = get_braces(each_ca, "bandParameterList-r13")
            if find_line2 > 0:
                result_dict["CA"].append(ca_combo_parse(picked_ca, 38))


def get_supportedBandListENDC_r15(log, result_dict):
    picked, find_line = get_braces(log, "supportedBandListEN-DC-r15")

    if find_line > 0:
        for each_nr in re.findall('bandNR-r15 [0-9]+', picked):
            result_dict["en-DC-r15"].append({"nr": "n" + re.sub("bandNR-r15 ", "", each_nr), "log": each_nr})


def reset_feat_set_dict():
    global feat_set_dict
    feat_set_dict = {"featureSetsDL-r15": [], "featureSetsDL-PerCC-r15": [],
                     "featureSetsUL-r15": [], "featureSetsUL-PerCC-r15": []}


def get_featureSetsEUTRA_r15(log):
    global feat_set_dict

    picked, find_line = get_braces(log, "featureSetsDL-r15")
    if find_line > 0:
        each_result = []
        for each in get_braces_array(picked, 54):
            index_str, line = get_braces(each, "featureSetPerCC-ListDL-r15")
            if line > 0:
                feat_set_dict["featureSetsDL-r15"].append(re.findall("[0-9]+", index_str))

    picked, find_line = get_braces(log, "featureSetsDL-PerCC-r15")
    if find_line > 0:
        each_result = []
        for each in get_braces_array(picked, 54):
            m = re.search("supportedMIMO-CapabilityDL-MRDC-r15 (.+)", each)
            layer = m.group(1) if m else "oneLayer"
            feat_set_dict["featureSetsDL-PerCC-r15"].append({"layer": layer, "log_snip": pick_inside_brace(each)})

    picked, find_line = get_braces(log, "featureSetsUL-r15")
    if find_line > 0:
        each_result = []
        for each in get_braces_array(picked, 54):
            index_str, line = get_braces(each, "featureSetPerCC-ListUL-r15")
            if line > 0:
                feat_set_dict["featureSetsUL-r15"].append(re.findall("[0-9]+", index_str))

    picked, find_line = get_braces(log, "featureSetsUL-PerCC-r15")
    if find_line > 0:
        each_result = []
        for each in get_braces_array(picked, 54):
            m = re.search("supportedMIMO-CapabilityUL-MRDC-r15 (.+)", each)
            layer = m.group(1) if m else "oneLayer"
            m = re.search("ul-256QAM-r15 (.+)", each)
            ul_256qam = m.group(1) if m else ""
            feat_set_dict["featureSetsUL-PerCC-r15"].append(
                {"layer": layer, "ul_256qam": ul_256qam, "log_snip": pick_inside_brace(each)})


def get_supportedBandListNR(log, result_dict):
    picked_bandlist_NR, find_line = get_braces(log, "supportedBandListNR[^-]")
    # print("get_supportedBandListNR: ", picked_bandlist_NR, find_line)
    if find_line > 0:
        # // Split to each band (indent=6)
        for each_nr in get_braces_array(picked_bandlist_NR, 6):
            each_result = {}
            each_result["bandNR"] = get_num(each_nr, "bandNR")
            if 1 <= each_result["bandNR"] <= 95:
                # FR1.
                dl, ignore = get_braces(each_nr, "channelBWs-DL fr1 :")
                dl_str = delete_indent("channelBWs-DL fr1 :\n" + dl, 8)
                each_result["channelBWs-DL"] = dl_str

                m = re.search("scs-15kHz '([0-9]{8} [0-9]{2})'B", dl_str)
                each_result["DL_scs-15kHz"] = m.group(1) if m else "???????? ??"

                m = re.search("scs-30kHz '([0-9]{8} [0-9]{2})'B", dl_str)
                each_result["DL_scs-30kHz"] = m.group(1) if m else "???????? ??"

                m = re.search("scs-60kHz '([0-9]{8} [0-9]{2})'B", dl_str)
                each_result["DL_scs-60kHz"] = m.group(1) if m else "???????? ??"

                ul, ignore = get_braces(each_nr, "channelBWs-UL fr1 :")
                ul_str = delete_indent("channelBWs-UL fr1 :\n" + ul, 8)
                each_result["channelBWs-UL"] = ul_str

                m = re.search("scs-15kHz '([0-9]{8} [0-9]{2})'B", ul_str)
                each_result["UL_scs-15kHz"] = m.group(1) if m else "???????? ??"

                m = re.search("scs-30kHz '([0-9]{8} [0-9]{2})'B", ul_str)
                each_result["UL_scs-30kHz"] = m.group(1) if m else "???????? ??"

                m = re.search("scs-60kHz '([0-9]{8} [0-9]{2})'B", ul_str)
                each_result["UL_scs-60kHz"] = m.group(1) if m else "???????? ??"

            elif 257 <= each_result["bandNR"] <= 261:
                # FR2.
                dl, ignore = get_braces(each_nr, "channelBWs-DL fr2 :")
                dl_str = delete_indent("channelBWs-DL fr2 :\n" + dl, 8)
                each_result["channelBWs-DL"] = dl_str

                m = re.search("scs-60kHz '([0-9]{3})'B", dl_str)
                each_result["DL_scs-60kHz"] = m.group(1) if m else "???"

                m = re.search("scs-120kHz '([0-9]{3})'B", dl_str)
                each_result["DL_scs-120kHz"] = m.group(1) if m else "???"

                ul, ignore = get_braces(each_nr, "channelBWs-UL fr2 :")
                ul_str = delete_indent("channelBWs-UL fr2 :\n" + ul, 8)
                each_result["channelBWs-UL"] = ul_str

                m = re.search("scs-60kHz '([0-9]{3})'B", ul_str)
                each_result["UL_scs-60kHz"] = m.group(1) if m else "???"

                m = re.search("scs-120kHz '([0-9]{3})'B", ul_str)
                each_result["UL_scs-120kHz"] = m.group(1) if m else "???"

            else:
                print("Cannot get valid band in supportedBandListNR.")
                continue

            result_dict["supportedBandListNR"].append(each_result)


def get_supportedBandCombinationList(log, result_dict, isMRDC = None):
    picked_Combo, find_line = get_braces(log, "supportedBandCombinationList")
    if find_line > 0:
        result = []

        # Split to each band (indent=6)
        for each_combo in get_braces_array(picked_Combo, 6):
            each_result = {"bandList": [],
                           "featureSetCombination": get_num(each_combo, "featureSetCombination")}
            if "dynamicPowerSharingENDC supported" in each_combo:
                each_result["dynamicPowerSharingENDC"] = "supported"
            ul_lst = []
            dl_lst = []
            nr_dl_lst = []
            nr_ul_lst = []

            for each_band in get_braces_array(each_combo, 12):
                m = re.search("bandEUTRA ([0-9]+)", each_band)
                bandList_dict = {}
                if m:
                    int_band = int(m.group(1))
                    bandList_dict["bandEUTRA"] = int_band

                    m = re.search("ca-BandwidthClassDL-EUTRA ([a-zA-Z]+)", each_band)
                    if m:
                        dl_lst.append([int_band, m.group(1).upper()])
                        bandList_dict["ClassDL-EUTRA"] = m.group(1).upper()

                    m = re.search("ca-BandwidthClassUL-EUTRA ([a-zA-Z]+)", each_band)
                    if m:
                        ul_lst.append([int_band, m.group(1).upper()])
                        bandList_dict["ClassUL-EUTRA"] = m.group(1).upper()

                else:
                    m = re.search("bandNR ([0-9]+)", each_band)
                    if m:
                        int_nr = int(m.group(1))
                        bandList_dict["bandNR"] = int_nr

                        m = re.search("ca-BandwidthClassDL-NR ([a-zA-Z]+)", each_band)
                        if m:
                            nr_dl_lst.append([int_nr, m.group(1).upper()])
                            bandList_dict["ClassDL-NR"] = m.group(1).upper()

                        m = re.search("ca-BandwidthClassUL-NR ([a-zA-Z]+)", each_band)
                        if m:
                            nr_ul_lst.append([int_nr, m.group(1).upper()])
                            bandList_dict["ClassUL-NR"] = m.group(1).upper()

                    else:
                        print("Cannot get band or nr from supportedBandCombinationList, weird.")

                each_result["bandList"].append(bandList_dict)
            dl_lst.sort()
            ul_lst.sort()
            nr_dl_lst.sort()
            nr_ul_lst.sort()
            if isMRDC:
                comb_str = "DC_"
            else:
                comb_str = "CA_"
            comb_str += '-'.join([str(a[0]) + a[1] for a in dl_lst]) + "_" if len(dl_lst) > 0 else ""
            comb_str += '-'.join(['n' + str(a[0]) + a[1] for a in nr_dl_lst])
            each_result["Comb"] = comb_str

            ul_str = '-'.join([str(a[0]) + a[1] for a in ul_lst]) + "_" if len(ul_lst) > 0 else ""
            ul_str += '-'.join(['n' + str(a[0]) + a[1] for a in nr_ul_lst])
            each_result["UL"] = ul_str
            if isMRDC:
                result_dict["MR-DC"].append(each_result)
            else:
                result_dict["NRCA"].append(each_result)

def get_supportedBandCombinationListMrDc(log, result_dict):
    MRDC_Capability, find_line = get_braces(log, "value UE-MRDC-Capability")
    if find_line > 0:
        get_supportedBandCombinationList(MRDC_Capability, result_dict, True)

def get_supportedBandCombinationListNrCa(log, result_dict):
    NrCa_Capability, find_line = get_braces(log, "value UE-NR-Capability")
    if find_line > 0:
        get_supportedBandCombinationList(NrCa_Capability, result_dict, False)

def get_featureSets(log, result_dict):
    picked, find_line = get_braces(log, "featureSetsDownlink\n")
    if find_line > 0:
        result_dict["featureSets"]["featureSetListPerDownlinkCC"] = []
        for each in get_braces_array(picked, 6):
            index_str, line = get_braces(each, "featureSetListPerDownlinkCC")
            if line > 0:
                result_dict["featureSets"]["featureSetListPerDownlinkCC"].append(re.findall("[0-9]+", index_str))

    picked, find_line = get_braces(log, "featureSetsDownlinkPerCC\n")
    if find_line > 0:
        result_dict["featureSets"]["featureSetsDownlinkPerCC"] = []
        for each in get_braces_array(picked, 6):
            m = re.search("supportedSubcarrierSpacingDL kHz([0-9]+)", each)
            scs = m.group(1) + "kHz" if m else "???kHz"

            m = re.search("supportedBandwidthDL fr[12] : mhz([0-9]+)", each)
            bw = m.group(1) + "MHz" if m else "???kHz"

            m = re.search("maxNumberMIMO-LayersPDSCH ([a-zA-Z]+)", each)
            layer = m.group(1) if m else "???"

            m = re.search("supportedModulationOrderDL qam([0-9]+)", each)
            mod = m.group(1) + "QAM" if m else "???"

            result_dict["featureSets"]["featureSetsDownlinkPerCC"].append(
                dict(scs=scs, bw=bw, layer=layer, mod=mod, log_snip=pick_inside_brace(each)))

    picked, find_line = get_braces(log, "featureSetsUplink\n")
    if find_line > 0:
        result_dict["featureSets"]["featureSetListPerUplinkCC"] = []
        for each in get_braces_array(picked, 6):
            index_str, line = get_braces(each, "featureSetListPerUplinkCC")
            if line > 0:
                result_dict["featureSets"]["featureSetListPerUplinkCC"].append(re.findall("[0-9]+", index_str))

    picked, find_line = get_braces(log, "featureSetsUplinkPerCC\n")
    if find_line > 0:
        result_dict["featureSets"]["featureSetsUplinkPerCC"] = []
        for each in get_braces_array(picked, 6):
            m = re.search("supportedSubcarrierSpacingUL kHz([0-9]+)", each)
            scs = m.group(1) + "kHz" if m else "???kHz"

            m = re.search("supportedBandwidthUL fr[12] : mhz([0-9]+)", each)
            bw = m.group(1) + "MHz" if m else "???kHz"

            m = re.search("maxNumberMIMO-LayersCB-PUSCH ([a-zA-Z]+)", each)
            layer = m.group(1) if m else "???"

            m = re.search("supportedModulationOrderUL qam([0-9]+)", each)
            mod = m.group(1) + "QAM" if m else "???"

            log_snip = each
            log_snip = "\n".join(log_snip.split("\n")[2:-1])
            log_snip = delete_indent(log_snip, 8)
            result_dict["featureSets"]["featureSetsUplinkPerCC"].append(
                dict(scs=scs, bw=bw, layer=layer, mod=mod, log_snip=log_snip))


def get_featureSetCombinations(log, result_dict, isMrDc = None):
    global feat_set_dict
    picked_feat_set, find_line = get_braces(log, "featureSetCombinations")
    if find_line > 0:
        for each_feat in get_braces_array(picked_feat_set, 4):
            each_feat_list = []
            for each_rat in get_braces_array(each_feat, 6):
                each_dict = {"downlinkSetEUTRA": [], "uplinkSetEUTRA": [], "downlinkSetNR": [], "uplinkSetNR": []}
                for each_cc in re.findall("downlinkSetEUTRA [0-9]+", each_rat):
                    each_dict["downlinkSetEUTRA"].append(re.sub("downlinkSetEUTRA ", "", each_cc))

                for each_cc in re.findall("uplinkSetEUTRA [0-9]+", each_rat):
                    each_dict["uplinkSetEUTRA"].append(re.sub("uplinkSetEUTRA ", "", each_cc))

                for each_cc in re.findall("downlinkSetNR [0-9]+", each_rat):
                    each_dict["downlinkSetNR"].append(re.sub("downlinkSetNR ", "", each_cc))

                for each_cc in re.findall("uplinkSetNR [0-9]+", each_rat):
                    each_dict["uplinkSetNR"].append(re.sub("uplinkSetNR ", "", each_cc))

                each_feat_list.append(each_dict)
            if isMrDc:
                result_dict["featureSetCombinations_MrDc"].append(each_feat_list)
            else:
                result_dict["featureSetCombinations"].append(each_feat_list)

def get_featureSetCombinations_MrDc(log, result_dict):
    picked_feat_set, find_line = get_braces(log, "value UE-MRDC-Capability ")
    get_featureSetCombinations(picked_feat_set, result_dict, True)

def get_featureSetCombinations_NrCa(log, result_dict):
    picked_feat_set, find_line = get_braces(log, "value UE-NR-Capability")
    get_featureSetCombinations(picked_feat_set, result_dict, False)

def link_linkset(result_dict):
    global feat_set_dict

    if len(result_dict["featureSetCombinations"]) > 0 and len(result_dict["NRCA"]) > 0:
        result_dict["featureSets"].update(feat_set_dict)
        reset_feat_set_dict()


def make_txt_output(each_cap, result_dict, out_dir, stem):
    time_str = re.sub(" +", " ", each_cap["Time"])  # Replace continuous space to one space.
    time_str = re.sub("[ :\.]", "_", time_str)  # Replace special char to under bar.

    p_out = out_dir.joinpath(stem + time_str + ".txt")

    with p_out.open(mode='w') as f:
        if len(result_dict["supportedBandListEUTRA"]) > 0:
            f.write("[supportedBandListEUTRA]\n")
            f.write(f"{', '.join(result_dict['supportedBandListEUTRA'])}\n\n")

        if len(result_dict["supportedBandListUTRA-FDD"]) > 0:
            f.write("[supportedBandListUTRA-FDD]\n")
            f.write(f"{', '.join(result_dict['supportedBandListUTRA-FDD'])}\n\n")

        if len(result_dict["en-DC-r15"]) > 0:
            f.write("[supportedBandListEN-DC-r15]\n")
            f.write(f"{', '.join([a['nr'] for a in result_dict['en-DC-r15']])}\n\n")

        if len(result_dict["CA"]) > 0:
            f.write("[supportedBandCombination]\n")
            for i, each_ca in enumerate(result_dict["CA"], 1):
                f.write(f"#{i:03} {each_ca['ca']}({each_ca['ca_layer']}) UL:{each_ca['ca_ul']}\n")
            f.write("\n\n")

        if len(result_dict["supportedBandListNR"]) > 0:
            f.write("[supportedBandListNR]\n")
            f.write(f"bandNR: {', '.join(['n' + str(a['bandNR']) for a in result_dict['supportedBandListNR']])}\n\n")

            for each_nr in result_dict["supportedBandListNR"]:
                bw_table_fr1 = ["5", "10", "15", "20", "25", "30", "40", "50", "60", "80"]
                bw_table_fr2 = ["50", "100", "200"]
                f.write(f"- BW for bandNR {each_nr['bandNR']}\n")
                for line in each_nr['channelBWs-DL'].split("\n"):
                    f.write(line)
                    support_bw = []
                    if " fr1 :" in each_nr['channelBWs-DL']:
                        m = re.search("([0-9]{8} *[0-9]{2})", line)
                        if m:
                            val = re.sub(" ", "", m.group(1))
                            if "1" in val:
                                for i, each_bit in enumerate(val):
                                    if each_bit == "1":
                                        support_bw.append(bw_table_fr1[i])
                                f.write(f"   => support BW {','.join(support_bw)} MHz")

                    elif " fr2 :" in each_nr['channelBWs-DL']:
                        m = re.search("\'([0-9]{3})\'", line)
                        if m and "1" in m.group(1):
                            for i, each_bit in enumerate(m.group(1)):
                                if each_bit == "1":
                                    support_bw.append(bw_table_fr2[i])
                            f.write(f"   => support BW {','.join(support_bw)} MHz")
                    f.write("\n")

                for line in each_nr['channelBWs-UL'].split("\n"):
                    f.write(line)
                    support_bw = []
                    if " fr1 :" in each_nr['channelBWs-UL']:
                        m = re.search("([0-9]{8} *[0-9]{2})", line)
                        if m:
                            val = re.sub(" ", "", m.group(1))
                            if "1" in val:
                                for i, each_bit in enumerate(val):
                                    if each_bit == "1":
                                        support_bw.append(bw_table_fr1[i])
                                f.write(f"   => support BW {','.join(support_bw)} MHz")
                    elif " fr2 :" in each_nr['channelBWs-UL']:
                        m = re.search("\'([0-9]{3})\'", line)
                        if m and "1" in m.group(1):
                            for i, each_bit in enumerate(m.group(1)):
                                if each_bit == "1":
                                    support_bw.append(bw_table_fr2[i])
                            f.write(f"   => support BW {','.join(support_bw)} MHz")
                    f.write("\n")
                f.write("\n\n")

        if len(result_dict["NRCA"]) > 0:
            set_endc_mrdc("NRCA", result_dict, f)

        f.write("\n\n")

        if len(result_dict['MR-DC']) > 0:
            set_endc_mrdc("MR-DC", result_dict, f)
            # print(1)
        f.write("\n---\n")

def set_endc_mrdc(type, result_dict, f):
    if len(result_dict[type]) > 0:
        if "MR-DC" == type:
            f.write("[supportedBandCombinationList_" + "ENDC" + "]\n")
        else:
            f.write("[supportedBandCombinationList_" + type + "]\n")

        if len(result_dict["featureSets"].get("featureSetsDL-r15", [])) > 0:
            for i, each_endc in enumerate(result_dict[type], 1):
                if type == "MR-DC":
                    f_set = result_dict["featureSetCombinations_MrDc"][each_endc["featureSetCombination"]]
                else:
                    f_set = result_dict["featureSetCombinations"][each_endc["featureSetCombination"]]
                num_cc_lst = []
                for each in f_set:
                    num_cc_lst.append(len(each.get("downlinkSetEUTRA", 0)))
                    num_cc_lst.append(len(each.get("downlinkSetNR", 0)))
                    num_cc_lst.append(len(each.get("uplinkSetEUTRA", 0)))
                    num_cc_lst.append(len(each.get("uplinkSetNR", 0)))

                for j in range(max(num_cc_lst)):
                    eutra_lst = []
                    nr_lst = []
                    code_lst = ""
                    conlunece_lst = ""
                    re_log = ""
                    for k, b_list in enumerate(each_endc["bandList"]):
                        if "ClassDL-EUTRA" in b_list:
                            if k > 0:
                                code_lst += "+"
                                conlunece_lst += "-"
                            cc_index = int(f_set[k]["downlinkSetEUTRA"][j])
                            featureSetDLPerCCIdList = result_dict["featureSets"]["featureSetsDL-r15"][cc_index - 1]
                            layer_str = ""
                            for index in featureSetDLPerCCIdList:
                                per_index = int(index)
                                layer_str = getLayerBw(per_index, "featureSetsDL-PerCC-r15", result_dict,layer_str)
                            #eutra_lst.append(f"{b_list['bandEUTRA']}{b_list['ClassDL-EUTRA']}({layer_str[:1]})")
                            eutra_lst.append(f"{b_list['bandEUTRA']}{b_list['ClassDL-EUTRA']}")
                            code_lst += f"B{b_list['bandEUTRA']}{b_list['ClassDL-EUTRA']}[{layer_str[:-1]}]"
                            conlunece_lst += f"b{b_list['bandEUTRA']}{b_list['ClassDL-EUTRA']}[{layer_str[:-1]}]"
                            re_log += f"b{b_list['bandEUTRA']}{b_list['ClassDL-EUTRA']}(.*)"
                            if "ClassUL-EUTRA" in b_list:
                                cc_index = int(f_set[k]["uplinkSetEUTRA"][j])
                                featureSetUlPerCCIdList = result_dict["featureSets"]["featureSetsUL-r15"][cc_index - 1]
                                layer_str = ""
                                for index in featureSetUlPerCCIdList:
                                    per_index = int(index)
                                    layer_str = getLayerBw(per_index,"featureSetsUL-PerCC-r15", result_dict, layer_str)
                                    code_lst += f";{b_list['ClassUL-EUTRA']}[{layer_str[:-1]}]"
                                    conlunece_lst +=  f"{b_list['ClassUL-EUTRA']}[{layer_str[:-1]}]"
                                    re_log += f"{b_list['ClassUL-EUTRA']}(.*)"
                        elif "ClassDL-NR" in b_list:
                            if len(code_lst) > 1:
                                code_lst += "+"
                            if len(conlunece_lst) >1:
                                conlunece_lst += "-"
                            cc_index = int(f_set[k]["downlinkSetNR"][j])
                            featureSetDLPerCCIdNrList = result_dict["featureSets"]["featureSetListPerDownlinkCC"][cc_index - 1]
                            layer_str = ""
                            layerCon_str = ""
                            for index in featureSetDLPerCCIdNrList:
                                per_index = int(index)
                                layer_str = getLayerBw(per_index, "featureSetsDownlinkPerCC", result_dict, layer_str, 1)
                                layerCon_str = getLayerBw(per_index, "featureSetsDownlinkPerCC", result_dict,
                                                          layerCon_str, 2)
                            nr_lst.append(f"n{b_list['bandNR']}{b_list['ClassDL-NR']}")
                            DL_Nr_list = f"N{b_list['bandNR']}{b_list['ClassDL-NR']}[{layer_str[:-1]}]"
                            code_lst += DL_Nr_list
                            conlunece_lst += f"n{b_list['bandNR']}{b_list['ClassDL-NR']}[{layerCon_str[:-1]}]"
                            re_log += f"n{b_list['bandNR']}{b_list['ClassDL-NR']}(.*)"
                            if "ClassUL-NR" in b_list:
                                featureSetULPerCCIdNrList = result_dict["featureSets"]["featureSetListPerUplinkCC"][cc_index - 1]
                                layer_str = ""
                                layerCon_str = ""
                                for index in featureSetULPerCCIdNrList:
                                    per_index = int(index)
                                    layer_str = getLayerBw(per_index, "featureSetsUplinkPerCC", result_dict, layer_str, 1)
                                    layerCon_str = getLayerBw(per_index, "featureSetsUplinkPerCC", result_dict, layerCon_str, 2)
                                UL_Nr_List = f"{b_list['ClassUL-NR']}[{layer_str[:-1]}]"
                                code_lst += ";" + UL_Nr_List
                                conlunece_lst +=  f"{b_list['ClassUL-NR']}[{layerCon_str[:-1]}]"
                                re_log += f"{b_list['ClassUL-NR']}(.*)"
                    if type == "MR-DC":
                        endc_str = "DC_"
                    elif type == "NRCA":
                        endc_str = "CA_"
                    if len(eutra_lst) > 0:
                        endc_str += "-".join(eutra_lst) + "_"
                    endc_str += "-".join(nr_lst)
                    if j == 0:
                        f.write("{:03} {} UL:{:<20}{:<120}{:>20}\n".format(i, endc_str, each_endc['UL'], code_lst, conlunece_lst))
                    else:
                        f.write(f"     {endc_str} UL:{each_endc['UL']}   --->  {code_lst}  --->  {conlunece_lst}\n")
                    print(f"#{i:03}  ", re_log)
        else:
            # No featureSetsDL-r15 info is present, skip to print layer info.
            for i, each_endc in enumerate(result_dict[type], 1):
                f.write(f"#{i-1:03} {each_endc['Comb']} UL:{each_endc['UL']}\n")

def getLayerBw(per_index, featureSet, result_dict, layer_str = "", isNr = 0):
    if isNr == 1:
        layer = result_dict["featureSets"][featureSet][per_index -1]["layer"]
        bw = re.search("[0-9]+", result_dict["featureSets"][featureSet][per_index - 1]["bw"]).group()
        if layer == "twoLayers":
            layer_str += bw + "x2,"
        elif layer == "fourLayers":
            layer_str += bw + "x4,"
        elif layer == "oneLayer":
            layer_str += bw + "x1,"
        return layer_str
    elif isNr == 2:
        layer = result_dict["featureSets"][featureSet][per_index -1]["layer"]
        bw = re.search("[0-9]+", result_dict["featureSets"][featureSet][per_index - 1]["bw"]).group()
        if layer == "twoLayers":
            layer_str += "2:"+ bw + ","
        elif layer == "fourLayers":
            layer_str +=  "4:"+ bw + ","
        elif layer == "oneLayer":
            layer_str +=  "1:"+ bw + ","
        return layer_str
    else:
        layer = result_dict["featureSets"][featureSet][per_index]["layer"]
        if layer == "twoLayers":
            layer_str += "2,"
        elif layer == "fourLayers":
            layer_str += "4,"
        elif layer == "oneLayer":
            layer_str += "1,"
        return layer_str

def make_excel_output(each_cap, result_dict, out_dir, stem):
    import openpyxl

    if len(result_dict["NRCA"]) == 0 and len(result_dict["CA"]) == 0:
        return

    time_str = re.sub(" +", " ", each_cap["Time"])  # Replace continuous space to one space.
    time_str = re.sub("[ :\.]", "_", time_str)  # Replace special char to under bar.

    p_out = out_dir.joinpath(stem + time_str + ".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active

    if len(result_dict["NRCA"]) > 0 or len(result_dict["MR-DC"]) > 0:
        if ws.title == "Sheet":
            ws.title = "NRCA ENDC Comb Capability"
        else:
            ws = wb.create_sheet(title="NRCA ENDC Comb Capability")

        # // Header
        ws.merge_cells("A1:A3")
        ws["A1"].value = "#"

        ws.merge_cells("B1:B3")
        ws["B1"].value = "DL Configration"

        ws.merge_cells("C1:C3")
        ws["C1"].value = "Uplink"

        ws.merge_cells("D1:I1")
        ws["D1"].value = "supportedBandCombinationList"

        ws.merge_cells("J1:O1")
        ws["J1"].value = "featureSetCombinations"

        ws.merge_cells("D2:D3")
        ws["D2"].value = "FeatureSetCombination-Id"

        ws.merge_cells("E2:E3")
        ws["E2"].value = "RAT"

        ws.merge_cells("F2:F3")
        ws["F2"].value = "Band"

        ws.merge_cells("G2:H2")
        ws["G2"].value = "ca-BandwidthClass"

        ws["I2"].value = "mrdc-Parameters"

        ws["J2"].value = "downlinkSet"

        ws["K2"].value = "uplinkSet"

        ws.merge_cells("L2:M2")
        ws["L2"].value = "DownlinkSet"

        ws.merge_cells("N2:O2")
        ws["N2"].value = "UplinkSet"

        ws["G3"].value = "Downlink"
        ws["H3"].value = "Uplink"
        ws["I3"].value = "dynamicPowerSharing"
        ws["J3"].value = "Downlink Set Id"
        ws["K3"].value = "Uplink Set Id"
        ws["L3"].value = "featureSetListPerDownlinkCC-Id"
        ws["M3"].value = "featureSetsDownlinkPerCC"
        ws["N3"].value = "featureSetListPerUplinkCC-Id"
        ws["O3"].value = "featureSetsUplinkPerCC"

        cur_row = 4
        if len(result_dict["NRCA"]) > 0:
            cur_row = set_ENDC_NRCA_excel(ws, result_dict["NRCA"], result_dict, cur_row, False)
        if len(result_dict["MR-DC"]) > 0:
            if cur_row > 4:
                cur_row += 2
            cur_row = set_ENDC_NRCA_excel(ws, result_dict["MR-DC"],result_dict, cur_row, True)

        # // Create separate sheet to record log snipped for EN-DC. (It is required to fill a report for KDDI)
        ws = wb.create_sheet(title="Comb Log Snip")

        # // Header
        ws["A1"].value = "Params"
        ws["B1"].value = "CC Index"
        ws["C1"].value = "Log Snipped"

        cur_row = 2

        for i, each in enumerate(result_dict["featureSets"]["featureSetsDL-PerCC-r15"]):
            if i == 0:
                ws.cell(cur_row, 1).value = "featureSetsDL-PerCC-r15"
            ws.cell(cur_row, 2).value = i
            ws.cell(cur_row, 3).value = each["log_snip"]
            cur_row += 1

        for i, each in enumerate(result_dict["featureSets"]["featureSetsUL-PerCC-r15"]):
            if i == 0:
                ws.cell(cur_row, 1).value = "featureSetsUL-PerCC-r15"
            ws.cell(cur_row, 2).value = i
            ws.cell(cur_row, 3).value = each["log_snip"]
            cur_row += 1

        for i, each in enumerate(result_dict["featureSets"]["featureSetsDownlinkPerCC"]):
            if i == 0:
                ws.cell(cur_row, 1).value = "featureSetsDownlinkPerCC"
            ws.cell(cur_row, 2).value = i + 1
            ws.cell(cur_row, 3).value = each["log_snip"]
            cur_row += 1

        for i, each in enumerate(result_dict["featureSets"]["featureSetsUplinkPerCC"]):
            if i == 0:
                ws.cell(cur_row, 1).value = "featureSetsUplinkPerCC"
            ws.cell(cur_row, 2).value = i + 1
            ws.cell(cur_row, 3).value = each["log_snip"]
            cur_row += 1

    if len(result_dict["CA"]) > 0:
        if ws.title == "Sheet":
            ws.title = "LTE CA Comb"
        else:
            ws = wb.create_sheet(title="LTE CA Comb")

        # // Header
        ws["A1"].value = "#"
        ws["B1"].value = "CA Comb(Downlink)"
        ws["C1"].value = "Layer(Donwlink)"
        ws["D1"].value = "Uplink"
        ws["E1"].value = "Log snip"

        for i, ca in enumerate(result_dict["CA"]):
            ws.cell(i + 2, 1).value = i + 1
            ws.cell(i + 2, 2).value = ca["ca"]
            ws.cell(i + 2, 3).value = ca["ca_layer"]
            ws.cell(i + 2, 4).value = ca["ca_ul"]
            ws.cell(i + 2, 5).value = ca["log_snip"]

    wb.save(p_out)

def set_ENDC_NRCA_excel(ws, CA_dict, result_dict, cur_row, isMrDC = False):
    for i, each in enumerate(CA_dict):
        if cur_row == 5:
            for j in [1, 2, 3, 4, 9]:
                ws.merge_cells(start_row=cur_row, start_column=j, end_row=cur_row + len(each["bandList"]) - 1,
                               end_column=j)
        if isMrDC:
            ws.cell(cur_row, 1).value = "{}_{}".format("ENDC", i + 1)
        else:
            ws.cell(cur_row, 1).value = "{}_{}".format("NRCA", i + 1)
        ws.cell(cur_row, 2).value = each["Comb"]
        ws.cell(cur_row, 3).value = each["UL"]
        ws.cell(cur_row, 4).value = each["featureSetCombination"]
        ws.cell(cur_row, 9).value = each.get("dynamicPowerSharingENDC", "-")
        if isMrDC == False:
            f_set_comb = result_dict["featureSetCombinations"][each["featureSetCombination"]]
        else:
            f_set_comb = result_dict["featureSetCombinations_MrDc"][each["featureSetCombination"]]
        f_set = result_dict["featureSets"]
        for j, each_band in enumerate(each["bandList"]):
            rat = "eutra" if "ClassDL-EUTRA" in each_band else "nr"
            ws.cell(cur_row + j, 5).value = rat
            if rat == "eutra":
                ws.cell(cur_row + j, 6).value = each_band.get("bandEUTRA")
                ws.cell(cur_row + j, 7).value = each_band.get("ClassDL-EUTRA").lower()
                ws.cell(cur_row + j, 8).value = each_band.get("ClassUL-EUTRA", "-").lower()
                ws.cell(cur_row + j, 10).value = "|".join(f_set_comb[j].get("downlinkSetEUTRA", ["-"]))
                ws.cell(cur_row + j, 11).value = "|".join(f_set_comb[j].get("uplinkSetEUTRA", ["-"]))

                if len(f_set["featureSetsDL-r15"]) > 0:
                    per_cc_lst = []
                    for id in f_set_comb[j]["downlinkSetEUTRA"]:
                        index = int(id) - 1
                        per_cc_lst.append(",".join(f_set["featureSetsDL-r15"][index]))
                    ws.cell(cur_row + j, 12).value = "|".join(per_cc_lst)

                    first_cc = per_cc_lst[0]
                    first_cc_int = int(first_cc.split(",")[0])

                    if "featureSetsDL-PerCC-r15" in f_set:
                        ws.cell(cur_row + j, 13).value = f_set["featureSetsDL-PerCC-r15"][first_cc_int]["layer"]

                if not "ClassUL-EUTRA" in each_band:
                    ws.cell(cur_row + j, 14).value = "-"
                    ws.cell(cur_row + j, 15).value = "-"

                elif len(f_set["featureSetsUL-r15"]) > 0:
                    per_cc_lst = []
                    for id in f_set_comb[j]["uplinkSetEUTRA"]:
                        index = int(id) - 1
                        per_cc_lst.append(",".join(f_set["featureSetsUL-r15"][index]))
                    ws.cell(cur_row + j, 14).value = "|".join(per_cc_lst)

                    first_cc = per_cc_lst[0]
                    first_cc_int = int(first_cc.split(",")[0])
                    if "featureSetsUL-PerCC-r15" in f_set:
                        ws.cell(cur_row + j, 15).value = f_set["featureSetsUL-PerCC-r15"][first_cc_int]["layer"]

            elif rat == "nr":
                ws.cell(cur_row + j, 6).value = each_band.get("bandNR")
                ws.cell(cur_row + j, 7).value = each_band.get("ClassDL-NR").lower()
                ws.cell(cur_row + j, 8).value = each_band.get("ClassUL-NR", "-").lower()
                ws.cell(cur_row + j, 10).value = "|".join(f_set_comb[j].get("downlinkSetNR", ["-"]))
                ws.cell(cur_row + j, 11).value = "|".join(f_set_comb[j].get("uplinkSetNR", ["-"]))

                if "featureSetListPerDownlinkCC" in f_set:
                    per_cc_lst = []
                    for id in f_set_comb[j]["downlinkSetNR"]:
                        index = int(id) - 1
                        per_cc_lst.append(",".join(f_set["featureSetListPerDownlinkCC"][index]))
                    ws.cell(cur_row + j, 12).value = "|".join(per_cc_lst)

                    # first_cc = per_cc_lst[0]
                    # first_cc_int = int(first_cc.split(",")[0])
                    # if "featureSetsDownlinkPerCC" in f_set:
                    #     tmp_cc = f_set["featureSetsDownlinkPerCC"][first_cc_int - 1]
                    #     tmp_cc_str = f"SCS:{tmp_cc['scs']} BW:{tmp_cc['bw']} {tmp_cc['layer']} {tmp_cc['mod']}"
                    #     tmp_cc_str = f"{tmp_cc['log_snip']}" # 返回整个featureSetsDownlinkPerCC
                    #     ws.cell(cur_row + j, 13).value = tmp_cc_str
                    tmp_cc_lst = [] # 返回整个featureSetsDownlinkPerCC,并添加分隔行
                    for per_cc in per_cc_lst:
                        per_cc_int = int(per_cc.split(",")[0])
                        if "featureSetsDownlinkPerCC" in f_set:
                            tmp_cc = f_set["featureSetsDownlinkPerCC"][per_cc_int - 1]
                            # tmp_cc_str = f"SCS:{tmp_cc['scs']} BW:{tmp_cc['bw']} {tmp_cc['layer']} {tmp_cc['mod']}"
                            tmp_cc_lst.append(f"{tmp_cc['log_snip']}")
                    ws.cell(cur_row + j, 13).value = "\n----------\n".join(tmp_cc for tmp_cc in tmp_cc_lst)

                if not "ClassUL-NR" in each_band:
                    ws.cell(cur_row + j, 14).value = "-"
                    ws.cell(cur_row + j, 15).value = "-"

                elif "featureSetListPerUplinkCC" in f_set:
                    per_cc_lst = []
                    for id in f_set_comb[j]["uplinkSetNR"]:
                        index = int(id) - 1
                        per_cc_lst.append(",".join(f_set["featureSetListPerUplinkCC"][index]))
                    ws.cell(cur_row + j, 14).value = "|".join(per_cc_lst)

                    # first_cc = per_cc_lst[0]
                    # first_cc_int = int(first_cc.split(",")[0])
                    # if "featureSetsUplinkPerCC" in f_set:
                    #     tmp_cc = f_set["featureSetsUplinkPerCC"][first_cc_int - 1]
                    #     tmp_cc_str = f"SCS:{tmp_cc['scs']} BW:{tmp_cc['bw']} {tmp_cc['layer']} {tmp_cc['mod']}"
                    #     tmp_cc_str = f"{tmp_cc['log_snip']}" # 返回整个featureSetsUplinkPerCC
                    #     ws.cell(cur_row + j, 15).value = tmp_cc_str
                    tmp_cc_lst = [] # 返回整个featureSetsUplinkPerCC,并添加分隔行
                    for per_cc in per_cc_lst:
                        per_cc_int = int(per_cc.split(",")[0])
                        if "featureSetsUplinkPerCC" in f_set:
                            tmp_cc = f_set["featureSetsUplinkPerCC"][per_cc_int - 1]
                            # tmp_cc_str = f"SCS:{tmp_cc['scs']} BW:{tmp_cc['bw']} {tmp_cc['layer']} {tmp_cc['mod']}"
                            tmp_cc_lst.append(f"{tmp_cc['log_snip']}")
                    ws.cell(cur_row + j, 15).value = "\n----------\n".join(tmp_cc for tmp_cc in tmp_cc_lst)

        cur_row += len(each["bandList"])
    return cur_row

def gather_from_qxdm_log(qxdm_log, stored_lst):
    import win32com.client

    # //Get QCAT Client
    try:
        QCAT = win32com.client.Dispatch('QCAT6.Application')
        QCAT.Visible = 0
    except:
        print("QCAT Application is required. Please install QCAT accordingly.")
        sys.exit()
    print(qxdm_log)
    print(dir(QCAT))
    QCAT.PacketFilter.SetAll(False)
    QCAT.PacketFilter.Set(0xB0C0, True)
    QCAT.PacketFilter.Set(0xB821, True)

    # //Filter Commitment
    QCAT.PacketFilter.Commit()

    if QCAT.OpenLog(qxdm_log) != 1:
        print("Open Log Error")

    # //Sort by timestamp.
    QCAT.SortByTime()

    # //Get log count
    allcnt = QCAT.PacketCount
    count = QCAT.VisiblePacketCount

    if count == 0:
        print("Seem no valid logs exist. Exit.")
        print("Please also check whether you put the file with fullpath format")
        QCAT.closeFile()
        sys.exit()

    log_pkt_store = []

    # //Get the initial log packet.
    log_packet = QCAT.FirstPacket
    print("count: " + str(count))
    # //Main Loop.
    for i in range(0, count):
        log_pkt_store.append(log_packet.Text)
        is_next = log_packet.next()
        if is_next == False:
            break
    # //End of Main Loop

    QCAT.closeFile()

    log = []
    index = 0
    collect_start = False

    for each_pkt in log_pkt_store:
        each_list = each_pkt.split("\r\n")
        for line in each_list:
            m = re.search(
                "^([0-9][0-9][0-9][0-9] +... +[0-9]+ +[0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9]) +\[..\] + ([^ ]+) +(.+$)",
                line)
            if m:
                collect_start = True
                log.append({"Time": m.group(1),
                            "Type": m.group(2),
                            "Description": m.group(3).rstrip(),
                            "cont": []})
                index = index + 1

            if collect_start:
                log[index - 1]["cont"].append(line.rstrip())

    enquiry = ""
    for each in log:
        print("log:", each)
        if each["Description"] == "LTE RRC OTA Packet  --  DL_DCCH / UECapabilityEnquiry" or \
                each["Description"] == "NR5G RRC OTA Packet  --  DL_DCCH / UeCapabilityEnquiry":
            enquiry = "\n".join(each["cont"])

        if each["Description"] == "LTE RRC OTA Packet  --  UL_DCCH / UECapabilityInformation" or \
                each["Description"] == "NR5G RRC OTA Packet  --  UL_DCCH / UeCapabilityInformation":
            each_dict = {"Time": each["Time"], "CapabilityLog": "\n".join(each["cont"]), "enquiry": enquiry}
            stored_lst.append(each_dict)
            enquiry = ""  # //Reset.


def gather_from_clipboard(stored_lst):
    import pyperclip
    current_clip = pyperclip.paste()  # Get data from clipboard.

    if isinstance(current_clip, str):
        collect_start = False
        log = []
        index = 0
        for line in current_clip.split('\n'):
            m = re.search(
                "^([0-9][0-9][0-9][0-9] +... +[0-9]+ +[0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9]) +\[..\] + ([^ ]+) +(.+$)",
                line)
            if m == None:
                m = re.search("^([0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9][0-9][0-9][0-9])\t+(\[.+\])\t+(.+$)", line)
            if m:
                collect_start = True
                log.append({"Time": m.group(1),
                            "Type": m.group(2),
                            "Description": m.group(3).rstrip(),
                            "cont": []})
                index = index + 1

            if collect_start:
                log[index - 1]["cont"].append(line.rstrip())

        enquiry = ""
        for each in log:
            if each["Description"] == "LTE RRC OTA Packet  --  DL_DCCH / UECapabilityEnquiry" or \
                    each["Description"] == "NR5G RRC OTA Packet  --  DL_DCCH / UeCapabilityEnquiry" or \
                    each["Description"] == "DL_DCCH / UeCapabilityEnquiry":
                enquiry = "\n".join(each["cont"])

            if each["Description"] == "LTE RRC OTA Packet  --  UL_DCCH / UECapabilityInformation" or \
                    each["Description"] == "NR5G RRC OTA Packet  --  UL_DCCH / UeCapabilityInformation" or \
                    each["Description"] == "UL_DCCH / UeCapabilityInformation":
                each_dict = {"Time": each["Time"], "CapabilityLog": "\n".join(each["cont"]), "enquiry": enquiry}
                stored_lst.append(each_dict)
                enquiry = ""  # //Reset.


def gather_from_txt_log(qcat_txt_log, stored_lst):
    log = []
    index = 0

    with open(qcat_txt_log, errors='replace') as f:
        collect_start = False
        for line in f.readlines():
            m = re.search(
                "^([0-9][0-9][0-9][0-9] +... +[0-9]+ +[0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9]) +\[..\] + ([^ ]+) +(.+$)",
                line)
            if m == None:
                m = re.search("^([0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9][0-9][0-9][0-9])\t+(\[.+\])\t+(.+$)", line)
            if m:
                collect_start = True
                log.append({"Time": m.group(1),
                            "Type": m.group(2),
                            "Description": m.group(3).rstrip(),
                            "cont": []})
                index = index + 1

            if collect_start:
                log[index - 1]["cont"].append(line.rstrip())
    # //Main Loop.
    enquiry = ""
    for each in log:
        if each["Description"] == "LTE RRC OTA Packet  --  DL_DCCH / UECapabilityEnquiry" or \
                each["Description"] == "NR5G RRC OTA Packet  --  DL_DCCH / UeCapabilityEnquiry" or \
                each["Description"] == "DL_DCCH / UeCapabilityEnquiry":
            enquiry = "\n".join(each["cont"])

        if each["Description"] == "LTE RRC OTA Packet  --  UL_DCCH / UECapabilityInformation" or \
                each["Description"] == "NR5G RRC OTA Packet  --  UL_DCCH / UeCapabilityInformation" or \
                each["Description"] == "UL_DCCH / UeCapabilityInformation":
            each_dict = {"Time": each["Time"], "CapabilityLog": "\n".join(each["cont"]), "enquiry": enquiry}
            stored_lst.append(each_dict)
            enquiry = ""  # //Reset.

class UeCapabilityCheckTool():
    def __init__(self, main_window:Tk):
        self.main_window = main_window
        self.log_stored_lst = []
        self.stem = "UE_"
        self.out_dir = Path(__file__).parent.parent / "result"
        self.out_dir.mkdir(parents=True, exist_ok=True)

    def set_init_window(self):
        self.main_window.title("UE能力检查工具")
        self.main_window.geometry('600x350')

        self.raw_data_label = Label(self.main_window, text='输入待分析的log: \n *应包含 DL_DCCH / UECapabilityEnquiry 和 UL_DCCH / UECapabilityInformation', anchor="w", justify="left")
        self.raw_data_label.pack(anchor='w',padx=10,pady=5)
        self.raw_data_frame = Frame(self.main_window)
        self.raw_data_frame.pack(expand=True, fill="both", padx=10, pady=10)
        scrollbar = ttk.Scrollbar(self.raw_data_frame, orient="vertical")
        self.raw_data_input = Text(self.raw_data_frame, height=10, wrap=NONE,  yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.raw_data_input.yview)
        scrollbar.pack(side="right", fill="y")
        self.raw_data_input.pack(side="left", expand=True, fill="both")
        self.raw_data_input.bind("<KeyRelease>", lambda event: self.raw_data_input.see("insert"))
        self.prefix_frame = Frame(self.main_window)
        self.prefix_frame.pack(anchor='w', padx=5, pady=5)
        self.prefix_label = Label(self.prefix_frame, text='结果文件名前缀:', anchor="w")
        self.prefix_label.pack(side=LEFT,padx=10,pady=5)
        default_prefix = StringVar()
        default_prefix.set("UECap")
        self.prefix_entry = Entry(self.prefix_frame, width=40, textvariable=default_prefix)
        self.prefix_entry.pack(side=LEFT,padx=5,pady=0)
        self.throughputs_chart_button = Button(self.main_window, text=' 生成结果(txt和excel) ', command=self.make_txt_output)
        self.throughputs_chart_button.pack(anchor='w',padx=10,pady=5)

    def gather_log(self):
        self.log_stored_lst = []
        log = []
        index = 0
        current_log = self.raw_data_input.get(1.0,END).splitlines()
        collect_start = False
        for line in current_log:
            m = re.search(
                "^([0-9][0-9][0-9][0-9] +... +[0-9]+ +[0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9]) +\[..\] + ([^ ]+) +(.+$)",
                line)
            if m == None:
                m = re.search("^([0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9][0-9][0-9][0-9])\t+(\[.+\])\t+(.+$)", line)
            if m == None:
                m = re.search("^([0-9][0-9]:[0-9][0-9]:[0-9][0-9]\.[0-9][0-9][0-9])\t+(\[.+\])\t+(.+$)", line)
            if m:
                collect_start = True
                log.append({"Time": m.group(1),
                            "Type": m.group(2),
                            "Description": m.group(3).rstrip(),
                            "cont": []})
                index = index + 1

            if collect_start:
                log[index - 1]["cont"].append(line.rstrip())
        # //Main Loop.
        enquiry = ""
        for each in log:
            if each["Description"] == "LTE RRC OTA Packet  --  DL_DCCH / UECapabilityEnquiry" or \
                    each["Description"] == "NR5G RRC OTA Packet  --  DL_DCCH / UeCapabilityEnquiry" or \
                    each["Description"] == "DL_DCCH / UECapabilityEnquiry" or \
                    each["Description"] == "DL_DCCH / UeCapabilityEnquiry":
                enquiry = "\n".join(each["cont"])

            if each["Description"] == "LTE RRC OTA Packet  --  UL_DCCH / UECapabilityInformation" or \
                    each["Description"] == "NR5G RRC OTA Packet  --  UL_DCCH / UeCapabilityInformation" or \
                    each["Description"] == "UL_DCCH / UECapabilityInformation" or \
                    each["Description"] == "UL_DCCH / UeCapabilityInformation":
                each_dict = {"Time": each["Time"], "CapabilityLog": "\n".join(each["cont"]), "enquiry": enquiry}
                self.log_stored_lst.append(each_dict)
                enquiry = ""  # //Reset.

    def make_txt_output(self):
        if self.prefix_entry.get():
            self.stem = self.prefix_entry.get() + "_"

        self.gather_log()
        if len(self.log_stored_lst) > 0:
            reset_feat_set_dict()

            for each_cap in self.log_stored_lst:
                result_dict = {"supportedBandListEUTRA": [], "supportedBandListUTRA-FDD": [], "en-DC-r15": [],
                            "featureSets": {}, "CA": [], "supportedBandListNR": [], "NRCA": [],
                            "featureSetCombinations": [], "featureSetCombinations_MrDc" : [],"MR-DC" : [] }
                get_supportedBandListEUTRA(each_cap["CapabilityLog"], result_dict)
                get_supportedBandListUTRA(each_cap["CapabilityLog"], result_dict)
                get_supportedBandCombination(each_cap["CapabilityLog"], result_dict)
                get_supportedBandListENDC_r15(each_cap["CapabilityLog"], result_dict)
                get_featureSetsEUTRA_r15(each_cap["CapabilityLog"])
                get_supportedBandListNR(each_cap["CapabilityLog"], result_dict)
                get_supportedBandCombinationListNrCa(each_cap["CapabilityLog"], result_dict)
                get_supportedBandCombinationListMrDc(each_cap["CapabilityLog"], result_dict)
                get_featureSets(each_cap["CapabilityLog"], result_dict)
                get_featureSetCombinations_NrCa(each_cap["CapabilityLog"], result_dict)
                get_featureSetCombinations_MrDc(each_cap["CapabilityLog"], result_dict)
                link_linkset(result_dict)
                make_txt_output(each_cap, result_dict, self.out_dir, self.stem)
                make_excel_output(each_cap, result_dict, self.out_dir, self.stem)
                # pp(result_dict)
            messagebox.showinfo("提示", f"已生成结果文件\n保存路径: {self.out_dir}\\")
        else:  # In case log_stored_lst equals 0,
            print("Cannot get valid UECapability Info.")

def start():
    init_window = Tk()
    # init_window.withdraw()
    ueCapabilityCheckTool = UeCapabilityCheckTool(init_window)
    ueCapabilityCheckTool.set_init_window()
    init_window.mainloop()

start()